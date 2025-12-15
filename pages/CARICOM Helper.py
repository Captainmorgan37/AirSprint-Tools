from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Sequence

import csv
from datetime import datetime, timezone
from functools import lru_cache

import streamlit as st
from openpyxl import load_workbook

from Home import configure_page, password_gate, render_sidebar
from fl3xx_api import (
    PassengerDetail,
    PreflightCrewMember,
    compute_fetch_dates,
    extract_crew_from_preflight,
    extract_passengers_from_pax_details,
    fetch_flights,
    fetch_flight_pax_details,
    fetch_preflight,
)
from flight_leg_utils import (
    FlightDataError,
    build_fl3xx_api_config,
    load_airport_metadata_lookup,
    normalize_fl3xx_payload,
    safe_parse_dt,
)
from zoneinfo_compat import ZoneInfo

configure_page(page_title="CARICOM Helper")
password_gate()
render_sidebar()

st.title("🌴 CARICOM Helper")

st.write(
    """
    Generate a CARICOM eAPIS workbook for an upcoming booking. Start by entering a
    **Booking Identifier** to pull the next 72 hours of flights from FL3XX. Any
    fields we cannot yet populate are left blank so the exported Excel template
    remains upload-ready while we build out the full data feed.
    """
)

CARICOM_COUNTRIES = {
    "Antigua and Barbuda",
    "Bahamas",
    "Barbados",
    "Belize",
    "Dominica",
    "Grenada",
    "Guyana",
    "Haiti",
    "Jamaica",
    "Montserrat",
    "Saint Kitts and Nevis",
    "Saint Lucia",
    "Saint Vincent and the Grenadines",
    "Suriname",
    "Trinidad and Tobago",
}

CARICOM_REQUIREMENT_COUNTRIES = {
    "antigua and barbuda",
    "barbados",
    "dominica",
    "grenada",
    "guyana",
    "jamaica",
    "saint kitts and nevis",
    "saint lucia",
    "saint vincent and the grenadines",
    "trinidad and tobago",
}

CARICOM_COUNTRY_CODE_TO_NAME = {
    # ISO2
    "AG": "Antigua and Barbuda",
    "BB": "Barbados",
    "DM": "Dominica",
    "GD": "Grenada",
    "GY": "Guyana",
    "JM": "Jamaica",
    "KN": "Saint Kitts and Nevis",
    "LC": "Saint Lucia",
    "VC": "Saint Vincent and the Grenadines",
    "TT": "Trinidad and Tobago",
    # ISO3
    "ATG": "Antigua and Barbuda",
    "BRB": "Barbados",
    "DMA": "Dominica",
    "GRD": "Grenada",
    "GUY": "Guyana",
    "JAM": "Jamaica",
    "KNA": "Saint Kitts and Nevis",
    "LCA": "Saint Lucia",
    "VCT": "Saint Vincent and the Grenadines",
    "TTO": "Trinidad and Tobago",
}


@st.cache_data(ttl=300, show_spinner=False)
def _fetch_upcoming_flights(settings: Dict[str, Any]) -> list[dict[str, Any]]:
    config = build_fl3xx_api_config(settings)
    from_date, to_date = compute_fetch_dates(inclusive_days=2)
    flights, _ = fetch_flights(config, from_date=from_date, to_date=to_date)
    return flights


@st.cache_data(ttl=300, show_spinner=False)
def _fetch_caricom_requirement_flights(settings: Dict[str, Any]) -> list[dict[str, Any]]:
    config = build_fl3xx_api_config(settings)
    from_date, to_date = compute_fetch_dates(inclusive_days=3)
    flights, _ = fetch_flights(config, from_date=from_date, to_date=to_date)
    return flights


def _normalise_legs(payload: Iterable[Any]) -> list[dict[str, Any]]:
    legs, _ = normalize_fl3xx_payload(list(payload))
    return legs


def _caricom_applicable_legs(legs: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    applicable: list[dict[str, Any]] = []
    for leg in legs:
        dep_airport = leg.get("departure_airport")
        arr_airport = leg.get("arrival_airport")
        if _is_caricom_airport(dep_airport) or _is_caricom_airport(arr_airport):
            applicable.append(dict(leg))
    return applicable


def _matches_booking(leg: Mapping[str, Any], booking_identifier: str) -> bool:
    target = booking_identifier.strip().lower()
    if not target:
        return False

    for key in ("bookingIdentifier", "bookingReference", "bookingCode", "bookingId"):
        value = leg.get(key)
        if isinstance(value, str) and value.strip().lower() == target:
            return True
    return False


def _format_date_time(value: Any) -> tuple[str, str]:
    if value is None:
        return "", ""

    try:
        dt = safe_parse_dt(str(value))
    except Exception:
        return str(value), ""

    return dt.strftime("%Y/%m/%d"), dt.strftime("%H:%M")


def _format_epoch_date(value: Any) -> str:
    try:
        epoch = int(value)
    except (TypeError, ValueError):
        return ""

    try:
        dt = datetime.fromtimestamp(epoch / 1000, tz=timezone.utc)
    except Exception:
        return ""

    return dt.strftime("%Y/%m/%d")


def _format_gender(value: Any) -> str:
    if value is None:
        return ""
    cleaned = str(value).strip().upper()
    if cleaned.startswith("M"):
        return "M"
    if cleaned.startswith("F"):
        return "F"
    return cleaned


def _crew_rows(ws) -> list[int]:
    rows: list[int] = []
    for row in ws.iter_rows(min_col=2, max_col=2, min_row=19):
        cell = row[0]
        if isinstance(cell.value, int):
            rows.append(cell.row)
    return rows


def _populate_crew_sheet(
    workbook, crew_roster: Sequence[PreflightCrewMember], dep_airport: str, arr_airport: str
) -> None:
    crew_ws = workbook["Crew List"]
    dep_iata = _icao_to_iata(dep_airport)
    arr_iata = _icao_to_iata(arr_airport)
    target_rows = _crew_rows(crew_ws)

    fields = (
        ("last_name", 3),
        ("first_name", 6),
        ("middle_name", 9),
        ("nationality_iso3", 11),
        ("gender", 13),
        ("birth_date", 14),
        ("document_number", 16),
        ("document_issue_country_iso3", 18),
        ("document_expiration", 20),
    )

    for member, row in zip(crew_roster, target_rows):
        for field, column in fields:
            value = getattr(member, field)
            if field in {"birth_date", "document_expiration"}:
                value = _format_epoch_date(value)
            elif field == "gender":
                value = _format_gender(value)
            crew_ws.cell(row=row, column=column).value = value or ""

        crew_ws.cell(row=row, column=22).value = dep_iata
        crew_ws.cell(row=row, column=23).value = arr_iata
        crew_ws.cell(row=row, column=24).value = arr_iata

    empty_columns = (3, 6, 9, 11, 13, 14, 16, 18, 20, 22, 23, 24)
    for row in target_rows[len(crew_roster) :]:
        for column in empty_columns:
            crew_ws.cell(row=row, column=column).value = ""


def _passenger_rows(ws) -> list[int]:
    rows: list[int] = []
    for row in ws.iter_rows(min_col=2, max_col=2, min_row=19):
        cell = row[0]
        if isinstance(cell.value, int):
            rows.append(cell.row)
    return rows


def _populate_passenger_sheet(
    workbook, passengers: Sequence[PassengerDetail], dep_airport: str, arr_airport: str
) -> None:
    pax_ws = workbook["Passenger List"]
    dep_iata = _icao_to_iata(dep_airport)
    arr_iata = _icao_to_iata(arr_airport)
    target_rows = _passenger_rows(pax_ws)

    fields = (
        ("last_name", 3),
        ("first_name", 6),
        ("middle_name", 9),
        ("nationality_iso3", 11),
        ("gender", 13),
        ("birth_date", 14),
        ("document_number", 16),
        ("document_issue_country_iso3", 18),
        ("document_expiration", 20),
    )

    for passenger, row in zip(passengers, target_rows):
        for field, column in fields:
            value = getattr(passenger, field)
            if field in {"birth_date", "document_expiration"}:
                value = _format_epoch_date(value)
            elif field == "gender":
                value = _format_gender(value)
            pax_ws.cell(row=row, column=column).value = value or ""

        pax_ws.cell(row=row, column=22).value = dep_iata
        pax_ws.cell(row=row, column=23).value = arr_iata
        pax_ws.cell(row=row, column=24).value = arr_iata

    empty_columns = (3, 6, 9, 11, 13, 14, 16, 18, 20, 22, 23, 24)
    for row in target_rows[len(passengers) :]:
        for column in empty_columns:
            pax_ws.cell(row=row, column=column).value = ""


def _build_workbook(
    leg: dict[str, Any],
    crew_roster: Sequence[PreflightCrewMember],
    passenger_roster: Sequence[PassengerDetail],
    dep_airport: str,
    arr_airport: str,
    crew_count: int,
) -> BytesIO:
    workbook = load_workbook("docs/caricomformFlightsV6.xlsx")
    general_ws = workbook["General Information"]

    dep_date, dep_time = _format_date_time(leg.get("dep_time"))
    arr_date, arr_time = _format_date_time(leg.get("arrival_time"))

    flight_identifier = leg.get("aircraft_callsign") or leg.get("callsign") or _tail_to_callsign(
        leg.get("tail") or leg.get("aircraftName")
    )

    general_ws["B13"] = flight_identifier
    general_ws["E13"] = leg.get("tail") or leg.get("aircraftName") or ""
    general_ws["H13"] = len(passenger_roster) or leg.get("paxNumber") or ""
    general_ws["K13"] = len(crew_roster) or crew_count or ""

    general_ws["B18"] = dep_date
    general_ws["E18"] = dep_time
    general_ws["H18"] = _icao_to_iata(dep_airport)

    general_ws["L18"] = arr_date
    general_ws["O18"] = arr_time
    general_ws["R18"] = _icao_to_iata(arr_airport)

    general_ws["O13"] = "AIRSPRINT"
    general_ws["S13"] = "4032161699"
    general_ws["V13"] = "dispatch@airsprint.com"

    _populate_flight_details(workbook, leg, dep_airport, arr_airport, flight_identifier)
    _populate_crew_sheet(workbook, crew_roster, dep_airport, arr_airport)
    _populate_passenger_sheet(workbook, passenger_roster, dep_airport, arr_airport)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _tail_to_callsign(tail: Any) -> str:
    if tail is None:
        return ""

    tail_sanitized = str(tail).replace("-", "").upper()
    tail_map = {

        "CGASL": "ASP816",
        "CFASV": "ASP812",
        "CFLAS": "ASP820",
        "CFJAS": "ASP822",
        "CFASF": "ASP827",
        "CGASE": "ASP846",
        "CGASK": "ASP839",
        "CGXAS": "ASP826",
        "CGBAS": "ASP875",
        "CFSNY": "ASP858",
        "CFSYX": "ASP844",
        "CFSBR": "ASP814",
        "CFSRX": "ASP864",
        "CFSJR": "ASP877",
        "CFASQ": "ASP821",
        "CFSDO": "ASP836",
        "CFASP": "ASP519",
        "CFASR": "ASP524",
        "CFASW": "ASP503",
        "CFIAS": "ASP511",
        "CGASR": "ASP510",
        "CGZAS": "ASP508",
        "CFASY": "ASP489",
        "CGASW": "ASP554",
        "CGAAS": "ASP567",
        "CFNAS": "ASP473",
        "CGNAS": "ASP642",
        "CGFFS": "ASP595",
        "CFSFS": "ASP654",
        "CGFSX": "ASP609",
        "CFSFO": "ASP668",
        "CFSNP": "ASP675",
        "CFSQX": "ASP556",
        "CFSFP": "ASP686",
        "CFSEF": "ASP574",
        "CFSDN": "ASP548",
        "CGFSD": "ASP655",
        "CFSUP": "ASP653",
        "CFSRY": "ASP565",
        "CGFSJ": "ASP501",
        "CGIAS": "ASP531",
    }

    return tail_map.get(tail_sanitized, "")


def _combine_date_time(date_str: str, time_str: str) -> str:
    return " ".join(part for part in (date_str, time_str) if part)


def _populate_flight_details(
    workbook,
    leg: Mapping[str, Any],
    dep_airport: str,
    arr_airport: str,
    flight_identifier: str,
) -> None:
    dep_date, dep_time = _format_date_time(leg.get("dep_time"))
    arr_date, arr_time = _format_date_time(leg.get("arrival_time"))

    dep_dt = _combine_date_time(dep_date, dep_time)
    arr_dt = _combine_date_time(arr_date, arr_time)

    flight_id = flight_identifier or _tail_to_callsign(leg.get("tail") or leg.get("aircraftName"))
    aircraft_name = leg.get("tail") or leg.get("aircraftName") or ""

    for sheet_name in ("Crew List", "Passenger List"):
        ws = workbook[sheet_name]
        ws["D11"] = flight_id
        ws["G11"] = aircraft_name
        ws["L11"] = dep_dt
        ws["O11"] = _icao_to_iata(dep_airport)
        ws["R11"] = arr_dt
        ws["U11"] = _icao_to_iata(arr_airport)


@st.cache_data(show_spinner=False)
def _airport_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    path = Path("Airport TZ.txt")
    if not path.exists():
        return lookup

    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            icao = (row.get("icao") or row.get("ICAO") or "").strip().upper()
            iata = (row.get("iata") or row.get("IATA") or "").strip().upper()
            if icao and iata:
                lookup[icao] = iata
    return lookup


def _icao_to_iata(code: Any) -> str:
    if code is None:
        return ""

    icao = str(code).strip().upper()
    lookup = _airport_lookup()
    return lookup.get(icao, icao)


@lru_cache(maxsize=1)
def _airport_metadata_lookup() -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for code, metadata in load_airport_metadata_lookup().items():
        lookup[code.upper()] = metadata
    return lookup


def _canonical_country_name(name: str | None) -> str:
    if not name:
        return ""
    cleaned = name.replace("St.", "Saint").replace("St ", "Saint ").strip()
    return cleaned.lower()


def _airport_country(code: Any) -> str:
    if code is None:
        return ""
    airport_code = str(code).strip().upper()
    metadata = _airport_metadata_lookup().get(airport_code)
    if not metadata:
        return ""
    country = metadata.get("country")
    if not country:
        return ""

    country_cleaned = str(country).strip().upper()
    resolved = CARICOM_COUNTRY_CODE_TO_NAME.get(country_cleaned, country_cleaned)
    return resolved


def _airport_timezone(code: Any) -> str:
    if code is None:
        return ""
    airport_code = str(code).strip().upper()
    metadata = _airport_metadata_lookup().get(airport_code)
    if not metadata:
        return ""
    tz = metadata.get("tz")
    return str(tz).strip() if tz else ""


def _is_caricom_airport(code: Any) -> bool:
    country = _canonical_country_name(_airport_country(code))
    if not country:
        return False
    return country in CARICOM_REQUIREMENT_COUNTRIES


def _format_local_departure_date(dep_time: Any, dep_airport: Any) -> str:
    if dep_time is None:
        return ""
    try:
        dt = safe_parse_dt(str(dep_time))
    except Exception:
        return ""

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    tz_name = _airport_timezone(dep_airport)
    if tz_name:
        try:
            dt = dt.astimezone(ZoneInfo(tz_name))
        except Exception:
            pass

    return dt.strftime("%Y-%m-%d")


booking_identifier = st.text_input("Booking Identifier", placeholder="e.g. ASP-123456")
fetch_button = st.button("Fetch Flight & Prepare Workbook", type="primary")

try:
    api_settings = st.secrets.get("fl3xx_api")  # type: ignore[attr-defined]
except Exception:
    api_settings = None

if not api_settings:
    st.error(
        "FL3XX API credentials are missing. Add them under `fl3xx_api` in `.streamlit/secrets.toml`."
    )
else:
    if fetch_button:
        try:
            with st.spinner("Fetching upcoming flights from FL3XX…"):
                raw_flights = _fetch_upcoming_flights(dict(api_settings))
                legs = _normalise_legs(raw_flights)
        except FlightDataError as exc:
            st.error(str(exc))
        except Exception as exc:  # pragma: no cover - runtime fetch failures
            st.error(f"Flight retrieval failed: {exc}")
        else:
            if not legs:
                st.warning("No flights returned for the upcoming 72-hour window.")
            else:
                matched_legs = [leg for leg in legs if _matches_booking(leg, booking_identifier)]

                if not matched_legs:
                    st.warning("No matching booking was found in the next 72 hours.")
                else:
                    sorted_legs = sorted(
                        matched_legs,
                        key=lambda leg: leg.get("dep_time") or "",
                    )
                    selected_leg = sorted_legs[0]

                    dep_airport = selected_leg.get("departure_airport")
                    arr_airport = selected_leg.get("arrival_airport")
                    crew_members = (
                        selected_leg.get("crewMembers")
                        if isinstance(selected_leg.get("crewMembers"), list)
                        else []
                    )
                    crew_count = len(crew_members)
                    passenger_roster: list[PassengerDetail] = []
                    passenger_count = selected_leg.get("paxNumber")
                    caricom_route = [code for code in (dep_airport, arr_airport) if code]

                    try:
                        config = build_fl3xx_api_config(dict(api_settings))
                        flight_id = (
                            selected_leg.get("id")
                            or selected_leg.get("flightId")
                            or selected_leg.get("flight_id")
                        )
                        crew_roster: list[PreflightCrewMember] = []
                        if flight_id:
                            with st.spinner("Loading preflight crew roster…"):
                                preflight_payload = fetch_preflight(config, flight_id)
                            crew_roster = extract_crew_from_preflight(preflight_payload)
                            if crew_roster:
                                crew_count = len(crew_roster)

                            try:
                                with st.spinner("Loading passenger roster…"):
                                    pax_payload = fetch_flight_pax_details(config, flight_id)
                                passenger_roster = extract_passengers_from_pax_details(pax_payload)
                                if passenger_roster:
                                    passenger_count = len(passenger_roster)
                            except Exception as pax_exc:  # pragma: no cover - runtime fetch failures
                                st.warning(
                                    f"Unable to load passenger roster from pax_details: {pax_exc}"
                                )
                        else:
                            st.warning("Flight ID missing on selected leg; crew roster will be left blank.")
                    except Exception as exc:  # pragma: no cover - runtime fetch failures
                        crew_roster = []
                        passenger_roster = []
                        st.warning(f"Unable to load crew roster from preflight: {exc}")

                    st.subheader("Flight summary")
                    st.json(
                        {
                            "Booking": booking_identifier,
                            "Tail": selected_leg.get("tail"),
                            "Departure": dep_airport,
                            "Arrival": arr_airport,
                            "Passenger Count": passenger_count,
                            "Crew Count": crew_count,
                            "Route Airports": caricom_route,
                        },
                        expanded=False,
                    )

                    st.info("Crew and passenger details are pulled directly from FL3XX when available.")

                    workbook_bytes = _build_workbook(
                        selected_leg,
                        crew_roster,
                        passenger_roster,
                        dep_airport,
                        arr_airport,
                        crew_count,
                    )
                    file_label = f"CARICOM_{booking_identifier or 'booking'}.xlsx"

                    st.download_button(
                        "Download CARICOM Excel",
                        data=workbook_bytes,
                        file_name=file_label,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
    else:
        st.info("Enter a booking identifier and click **Fetch** to build a CARICOM workbook.")

st.divider()

st.subheader("CARICOM requirement scanner")
st.write(
    "Check upcoming flights for routes that depart from or arrive at CARICOM airports "
    "in the next three days."
)

scan_button = st.button("Scan upcoming routes for CARICOM requirements")

if not api_settings:
    st.info("Add FL3XX API credentials to enable CARICOM requirement scanning.")
elif scan_button:
    try:
        with st.spinner("Scanning upcoming flights for CARICOM routing…"):
            raw_flights = _fetch_caricom_requirement_flights(dict(api_settings))
            legs = _normalise_legs(raw_flights)
    except FlightDataError as exc:
        st.error(str(exc))
    except Exception as exc:  # pragma: no cover - runtime fetch failures
        st.error(f"Flight retrieval failed: {exc}")
    else:
        caricom_legs = _caricom_applicable_legs(legs)
        if not caricom_legs:
            st.success("No CARICOM routes found in the next three days.")
        else:
            rows = []
            for leg in sorted(caricom_legs, key=lambda item: item.get("dep_time") or ""):
                dep_airport = leg.get("departure_airport")
                arr_airport = leg.get("arrival_airport")
                dep_label = _icao_to_iata(dep_airport) or (dep_airport or "")
                arr_label = _icao_to_iata(arr_airport) or (arr_airport or "")
                routing = f"{dep_label or '—'} ➜ {arr_label or '—'}"
                rows.append(
                    {
                        "Routing": routing,
                        "Departure Date (Local)": _format_local_departure_date(
                            leg.get("dep_time"), dep_airport
                        ),
                        "Flight ID": leg.get("flightId")
                        or leg.get("flight_id")
                        or leg.get("id"),
                    }
                )

            st.dataframe(rows, hide_index=True)

